import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font

caminho_planilha = os.path.join(
    os.path.dirname(__file__), "..", "data", "MDIP_2025_tratado.xlsx"
)
df = pd.read_excel(caminho_planilha)

wb = load_workbook(caminho_planilha)
ws = wb.active

max_row = ws.max_row
max_col = ws.max_column
last_col_letter = ws.cell(row=1, column=max_col).column_letter
table_range = f"A1:{last_col_letter}{max_row}"

table = Table(displayName="Tabela", ref=table_range)

style = TableStyleInfo(
    name='TableStyleMedium2',
    showRowStripes=True
    )

table.tableStyleInfo = style
ws.add_table(table)

for cell in ws[1]:
    cell.font = Font(color="FFFFFF")

wb.save(caminho_planilha)

print(f"Arquivo XLSX formatado: {caminho_planilha}")
